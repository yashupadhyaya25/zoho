import pandas as pd
from openpyxl.comments import Comment
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import numpy as np

comment_df = pd.read_excel('./shadow_utilization.xlsx',sheet_name='Sheet1')
timesheet_df = pd.read_excel('./New Reports/Report_WO_Task.xlsx',sheet_name='Sheet1')
timesheet_df.columns = timesheet_df.columns.to_series().mask(lambda x: x.str.startswith('Unnamed')).ffill()
timesheet_df = timesheet_df.T.reset_index().T
unique_row_make = timesheet_df.iloc[0:2].T
excel_file = './New Reports/Report_WO_Task.xlsx'
wb = load_workbook(excel_file, data_only = True)
sh = wb['Sheet1']

unique_ls = []
for data in unique_row_make.itertuples():
    project_name = data[1]
    resource_name = data[2]
    # print(project_name,job_name,resource_name)
    unique_row = project_name + '-' + resource_name
    unique_ls.append(unique_row)
timesheet_df.loc[1] = unique_ls

comm_df = timesheet_df.iloc[2:3]

for data in comment_df.itertuples():
    date = data[1]
    resource_name = data[2]
    project_name = data[3]
    job_name = data[4]
    comment = data[5]
    row_inx_value = project_name + '-' + resource_name
    row_inx = 1
    col_inx = 1

    for data in timesheet_df.itertuples() :
        if data[1] == date :

            break
        row_inx += 1

    for data in comm_df.itertuples() :
        for resource in data[1:] :
            if row_inx_value == resource :
                break
            col_inx += 1

    sh.cell(row=row_inx, column=col_inx).fill = PatternFill(start_color='FFFFFF00',
                                                        end_color='FFFFFF00',
                                                        fill_type="solid")
    sh.cell(row=row_inx, column=col_inx).fill.start_color.index == 'FFFFFF00'
    sh.cell(row=row_inx, column=col_inx).comment = Comment(comment, "author name")

wb.save('./New Reports/Report_WO_Task_With_Shadow_Comment.xlsx')